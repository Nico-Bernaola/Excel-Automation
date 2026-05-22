import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


FUENTE = "Arial"

COLORES = {
    "header_bg":  "1F3864",
    "header_fg":  "FFFFFF",
    "fila_par":   "EEF2F7",
    "fila_impar": "FFFFFF",
    "titulo_bg":  "2E75B6",
    "titulo_fg":  "FFFFFF",
}


def format_and_save(state: dict, output_dir: Path) -> tuple:
    output_dir.mkdir(exist_ok=True)
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_base = f"{state['nombre_archivo']}_clean_{fecha}"

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_datos(wb, state["df_clean"])
    for nombre, df in state.get("analysis", {}).items():
        if isinstance(df, pd.DataFrame) and not df.empty:
            _sheet_resumen(wb, df, nombre.replace("_", " ").title())

    output_dir = output_dir / state["nombre_archivo"]
    output_dir.mkdir(exist_ok=True)
    xlsx_path = output_dir / f"{nombre_base}.xlsx"
    wb.save(xlsx_path)

    txt_path = None
    if state.get("insights"):
        txt_path = output_dir / f"{nombre_base}_insights.txt"
        txt_path.write_text(_construir_txt(state), encoding="utf-8")

    return xlsx_path, txt_path


def _construir_txt(state: dict) -> str:
    lineas = [
        f"INSIGHTS — {state['nombre_archivo']}",
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "=" * 60,
        "",
        state["insights"],
        "",
        "=" * 60,
        "REPORTE DE LIMPIEZA",
        "=" * 60,
    ]
    for _, msg, detalle in state["log"]:
        lineas.append(f"✓ {msg}")
        for d in detalle:
            lineas.append(f"    · {d}")
    return "\n".join(lineas)


# ── Sheets ────────────────────────────────────────────────────────────────────

def _sheet_datos(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Datos Limpios")
    cols_num = set(df.select_dtypes(include="number").columns)
    _titulo(ws, "Datos Limpios", len(df.columns))
    _headers(ws, df.columns, row=2)
    for i, row in enumerate(df.itertuples(index=False), start=3):
        par = i % 2 == 0
        for j, (col, val) in enumerate(zip(df.columns, row), start=1):
            cell = ws.cell(row=i, column=j, value=None if pd.isna(val) else val)
            _estilo_dato(cell, par, col in cols_num)
    _autofit(ws, df)
    ws.freeze_panes = "A3"


def _sheet_resumen(wb: Workbook, df: pd.DataFrame, titulo: str):
    ws = wb.create_sheet(titulo)
    cols_num = set(df.select_dtypes(include="number").columns)
    _titulo(ws, titulo, len(df.columns))
    _headers(ws, df.columns, row=2)
    for i, row in enumerate(df.itertuples(index=False), start=3):
        par = i % 2 == 0
        for j, (col, val) in enumerate(zip(df.columns, row), start=1):
            cell = ws.cell(row=i, column=j, value=None if pd.isna(val) else val)
            _estilo_dato(cell, par, col in cols_num)
    _autofit(ws, df)
    ws.freeze_panes = "A3"


# ── Estilos ───────────────────────────────────────────────────────────────────

def _titulo(ws, texto: str, n_cols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=texto)
    cell.font = Font(name=FUENTE, bold=True, size=12, color=COLORES["titulo_fg"])
    cell.fill = PatternFill("solid", fgColor=COLORES["titulo_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def _headers(ws, columnas, row: int):
    for j, col in enumerate(columnas, start=1):
        cell = ws.cell(row=row, column=j, value=str(col).replace("_", " ").title())
        cell.font = Font(name=FUENTE, bold=True, color=COLORES["header_fg"], size=10)
        cell.fill = PatternFill("solid", fgColor=COLORES["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _borde()
        ws.row_dimensions[row].height = 20


def _estilo_dato(cell, par: bool, es_numero: bool):
    color = COLORES["fila_par"] if par else COLORES["fila_impar"]
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name=FUENTE, size=10)
    cell.alignment = Alignment(horizontal="right" if es_numero else "left", vertical="center")
    cell.border = _borde()
    if es_numero and cell.value is not None:
        cell.number_format = "#,##0.00"


def _autofit(ws, df: pd.DataFrame):
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            df[col].dropna().astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 45)


def _borde():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)
