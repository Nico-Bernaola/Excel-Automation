import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


FONT = "Arial"

COLORS = {
    "header_bg":   "1F3864",
    "header_fg":   "FFFFFF",
    "even_row":    "EEF2F7",
    "odd_row":  "FFFFFF",
    "title_bg":   "2E75B6",
    "title_fg":   "FFFFFF",
    "anomaly":    "FFD966",  # amarillo para celdas con anomalías
}


def format_and_save(state: dict, output_dir: Path) -> tuple:
    output_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"{state['file_name']}_clean_{timestamp}"

    anomalies = state.get("anomalies", [])
    anomalous_cells = {(a["row"], a["column"]) for a in anomalies if a["row"] is not None}

    wb = Workbook()
    wb.remove(wb.active) # type: ignore

    _sheet_data(wb, state["df_clean"], anomalous_cells)

    for name, df in state.get("analysis", {}).items():
        if isinstance(df, pd.DataFrame) and not df.empty:
            _sheet_summary(wb, df, name.replace("_", " ").title())

    xlsx_path = output_dir / f"{base_name}.xlsx"
    wb.save(xlsx_path)

    txt_path = None
    if state.get("insights") or anomalies:
        txt_path = output_dir / f"{base_name}_insights.txt"
        txt_path.write_text(_build_txt(state), encoding="utf-8")

    return xlsx_path, txt_path


def _build_txt(state: dict) -> str:
    lines = [
        f"INSIGHTS — {state['file_name']}",
        f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "=" * 60,
        "",
    ]

    if state.get("insights"):
        lines += [state["insights"], ""]

    anomalies = state.get("anomalies", [])
    if anomalies:
        lines += ["=" * 60, f"ANOMALIES DETECTED ({len(anomalies)})", "=" * 60]
        for a in anomalies:
            lines.append(f"⚠ {a['message']}")
        lines.append("")

    from modules.history import format_comparison
    comparison_txt = format_comparison(state.get("comparison", {}))
    if comparison_txt:
        lines += ["", comparison_txt, ""]

    lines += ["=" * 60, "CLEANING REPORT", "=" * 60]
    for _, msg, detail in state["log"]:
        lines.append(f"✓ {msg}")
        for d in detail:
            lines.append(f"    · {d}")

    return "\n".join(lines)


# ── Sheets ────────────────────────────────────────────────────────────────────

def _sheet_data(wb: Workbook, df: pd.DataFrame, anomalous_cells: set):
    ws = wb.create_sheet("Clean Data")
    cols_num = set(df.select_dtypes(include="number").columns)
    cols_list = list(df.columns)

    _title(ws, "Clean Data", len(cols_list))
    _headers(ws, cols_list, row=2)

    for i, row in enumerate(df.itertuples(index=False), start=3):
        df_idx = i - 3
        pair = i % 2 == 0
        for j, (col, val) in enumerate(zip(cols_list, row), start=1):
            cell = ws.cell(row=i, column=j, value=None if pd.isna(val) else val)
            is_anomaly = (df_idx, col) in anomalous_cells
            _data_style(cell, pair, col in cols_num, is_anomaly)

    _autofit(ws, df)
    ws.freeze_panes = "A3"


def _sheet_summary(wb: Workbook, df: pd.DataFrame, title: str):
    ws = wb.create_sheet(title)
    cols_num = set(df.select_dtypes(include="number").columns)
    _title(ws, title, len(df.columns))
    _headers(ws, df.columns, row=2)
    for i, row in enumerate(df.itertuples(index=False), start=3):
        pair = i % 2 == 0
        for j, (col, val) in enumerate(zip(df.columns, row), start=1):
            cell = ws.cell(row=i, column=j, value=None if pd.isna(val) else val)
            _data_style(cell, pair, col in cols_num, is_anomaly=False)
    _autofit(ws, df)
    ws.freeze_panes = "A3"


# ── Estilos ───────────────────────────────────────────────────────────────────

def _title(ws, text: str, n_cols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = Font(name=FONT, bold=True, size=12, color=COLORS["title_fg"])
    cell.fill = PatternFill("solid", fgColor=COLORS["title_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def _headers(ws, columns, row: int):
    for j, col in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=j, value=str(col).replace("_", " ").title())
        cell.font = Font(name=FONT, bold=True, color=COLORS["header_fg"], size=10)
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws.row_dimensions[row].height = 20


def _data_style(cell, pair: bool, is_number: bool, is_anomaly: bool = False):
    if is_anomaly:
        color = COLORS["anomaly"]
    else:
        color = COLORS["even_row"] if pair else COLORS["odd_row"]
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name=FONT, size=10)
    cell.alignment = Alignment(horizontal="right" if is_number else "left", vertical="center")
    cell.border = _border()
    if is_number and cell.value is not None:
        cell.number_format = "#,##0.00"


def _autofit(ws, df: pd.DataFrame):
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            df[col].dropna().astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 45)


def _border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)
